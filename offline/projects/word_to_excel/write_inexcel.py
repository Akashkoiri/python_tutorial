import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\AKASH-PC\\Documents\\Python\\offline\\projects\\word_to_excel\\test.xlsx')

sh1 = wb['Sheet1']

sh1.cell(row=1,column=1,value='Akash')
sh1.cell(row=1,column=2,value='koiri')

wb.save('C:\\Users\\AKASH-PC\\Documents\\Python\\offline\\projects\\word_to_excel\\test.xlsx')